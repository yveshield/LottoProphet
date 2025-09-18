#!/usr/bin/python
#coding=utf-8

#author   : oahzxd
#version  : 1.0
#datetime : 2015.7.27

import os
import re
import sys
import time
import importlib
import functools
from copy import copy
import csv
import collections
import openpyxl
from itertools import combinations
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

import time
import functools

def timeit(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.perf_counter()
        result = func(*args, **kwargs)
        end = time.perf_counter()
        elapsed = end - start

        if elapsed < 1:
            print(f"{func.__name__} executed in {elapsed * 1000:.3f} ms")
        else:
            print(f"{func.__name__} executed in {elapsed:.3f} s")
        
        return result
    return wrapper


csv_file = r"scripts\ssq\ssq_history.csv"

fontRed = Font(color='FF0000', bold=True)
fontBold = Font(bold=True)

colors = [
    "FF0000", "0000FF", "008000", "FFA500", "800080", "000000", "808080", "00FFFF",
    "FFC0CB", "A52A2A", "FFD700", "808000", "00FF00", "FF00FF", "00CED1", "4169E1",
    "2F4F4F", "FF4500", "DA70D6", "7FFF00", "DC143C", "1E90FF", "B8860B", "008B8B",
    "B22222", "32CD32", "8B008B", "556B2F", "FF1493", "696969", "CD5C5C", "E9967A",
    "48D1CC", "C71585", "191970", "4682B4", "9ACD32", "F4A460", "D2691E"
]

redBalls = [[0 for col in range(6)] for row in range(500)]
blueBalls = [0 for row in range(500)]
omitTable = [[-1 for col in range(34)] for row in range(500)]

arrSeral = []
arrDate = []
startSeral = 0

"获取网络数据"
"sNo: 最近期数"
"sNo=0: 为输入截止日期方法查询"

@timeit
def getData():
    global arrSeral, arrDate, redBalls, blueBalls

    df = pd.read_csv(csv_file, encoding='gbk')
    df = df.head(100).iloc[::-1].reset_index(drop=True)

    # 重新初始化全局数组
    redBalls = [[0 for col in range(6)] for row in range(len(df))]
    blueBalls = [0 for row in range(len(df))]
    arrSeral.clear()
    arrDate.clear()

    # 遍历 CSV 记录
    offset = 0
    for index, row in df.iloc[::-1].iterrows():
        arrSeral.append(str(row["期数"]))  # 存储期号
        arrDate.append("")  # 原网页代码有日期，但 CSV 无日期数据，这里填充空字符串

        # 存储红球号码
        for col in range(6):
            redBalls[offset][col] = int(row[f"红球_{col+1}"])

        # 存储蓝球号码
        blueBalls[offset] = int(row["蓝球"])
        offset = offset+1

    setOmitTable()  # 调用原有的遗漏计算逻辑

@timeit
def convert_csv_to_formatted_excel():
    """
    将 CSV 转换为 Excel 并格式化：
    - 所有单元格水平垂直居中
    - 冻结第一行和第一列
    - 自动调整列宽
    """
    csv_path = r'scripts\ssq\ssq_history.csv'
    xlsx_path = r'scripts\ssq\ssq_history.xlsx'
    # 读取 CSV 内容
    with open(csv_path, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        rows = list(reader)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "双色球历史"

    # 写入数据并设置居中
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if 2 <= cell.column:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 红球列（第2列到第7列，索引从1开始）
            if 2 <= cell.column <= 7:
                cell.font = Font(color="FF0000")  # 红色
            # 蓝球列（第8列）
            elif cell.column == 8:
                cell.font = Font(color="0000FF")  # 蓝色

    # 冻结第一列和第一行
    ws.freeze_panes = 'B2'

    # 自动设置列宽
    for col_idx in range(1, ws.max_column + 1):
        max_length = max(len(str(ws.cell(row=r, column=col_idx).value)) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_length + 2, 20))

    # 保存 Excel
    wb.save(xlsx_path)

def setOmitTable():
    for i in range(1, 34): # 1-33号码
        omit = 0
        for j in range(99, -1, -1):
            for k in range(0, 6):
                if redBalls[j][k] == i:
                    omit = 0
            omitTable[j][i] = omit
            omit = omit + 1

@timeit
def analyze_lottery_repeats():
    # 记录红球的中奖期次
    red_ball_appearances = defaultdict(list)

    # 读取CSV文件
    with open(csv_file, newline='', encoding='gbk') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头

        for index, row in enumerate(reader, start=1):
            red_balls = map(int, row[1:7])  # 取红球号码
            for ball in red_balls:
                red_ball_appearances[ball].append(index)  # 记录红球出现的期次

    # 统计重复中奖次数
    repeat_counts = {ball: [0] * 9 for ball in range(1, 34)}

    for ball, periods in red_ball_appearances.items():
        gaps = []
        prev = None

        for p in periods:
            if prev is not None:
                if p - prev > 1:
                    gaps.append(1)  # 记录一次新的重复中奖
                elif gaps:  # 确保 gaps 至少有一个元素
                    gaps[-1] += 1
            prev = p

        # 统计各个重复中奖次数
        for gap in gaps:
            if gap <= 7:
                repeat_counts[ball][gap] += 1

    # 计算重复中奖比率
    for ball in range(1, 34):
        one_time = repeat_counts[ball][1]
        multi_time = sum(repeat_counts[ball][2:8])
        ratio = round(multi_time / one_time, 2) if one_time > 0 else 0.0
        repeat_counts[ball][8] = ratio

    # 打印结果
    print("红球 | 1次 | 2次 | 3次 | 4次 | 5次 | 6次 | 7次 | 比率")
    print("-" * 50)
    for ball in range(1, 34):
        print(f"{ball:>3}  | {repeat_counts[ball][1]:>3} | {repeat_counts[ball][2]:>3} | {repeat_counts[ball][3]:>3} |"
              f" {repeat_counts[ball][4]:>3} | {repeat_counts[ball][5]:>3} | {repeat_counts[ball][6]:>3} |"
              f" {repeat_counts[ball][7]:>3} | {repeat_counts[ball][8]:>4.2f}")


"1.奇偶数"
@timeit
def fileOddOrEven(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    oddOrEven(ws, 10)
    oddOrEven(ws, 5)
    oddOrEven(ws, 1)

def oddOrEven(ws, sNo):
    odd = 0
    even = 0
    for i in range(startSeral, startSeral + sNo):
        for j in range(0, 6):
            if (redBalls[i][j] % 2 == 0):
                even = even + 1
            else:
                odd = odd + 1

    if (sNo != 1):
        if (even > odd):
            ws.append([f"近{sNo}期有", f"奇 = {odd}", f"偶 = {even}", f"偶：+{(even * 100 // odd) - 100}%"])
        elif (even < odd):
            ws.append([f"近{sNo}期有", f"奇 = {odd}", f"偶 = {even}", f"奇：+{(odd * 100 // even) - 100}%"])
        else:
            ws.append([f"近{sNo}期有", f"奇 = {odd}", f"偶 = {even}", '奇 = 偶'])
    else:
        ws.append(['本期比例:', '奇:偶 = %d:%d' % (odd, even)])

"2.大小数"
@timeit
def fileBigOrSmall(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    bigOrSmall(ws, 10)
    bigOrSmall(ws, 5)
    bigOrSmall(ws, 1)

def bigOrSmall(ws, sNo):
    big = 0
    small = 0
    for i in range(startSeral, startSeral + sNo):
        for j in range(0, 6):
            if (redBalls[i][j] <= 16):
                small = small + 1
            else:
                big = big + 1

    if (sNo != 1):
        if (small > big):
            ws.append([f"近{sNo}期有", f"大 = {big}", f"小 = {small}", f"小：+{(small * 100 // big) - 100}%"])
        elif (small < big):
            ws.append([f"近{sNo}期有", f"大 = {big}", f"小 = {small}", f"大：+{(big * 100 // small) - 100}%"])
        else:
            ws.append([f"近{sNo}期有", f"大 = {big}", f"小 = {small}", '大 = 小'])
    else:
        ws.append(['本期比例:', '大:小 = %d:%d' % (big, small)])

# 缓存读取结果
_number_to_superscript = None
_superscript_map = {
    0: "⁰", 1: "¹", 2: "²", 3: "³", 4: "⁴", 5: "⁵"
}

@timeit
def load_superscript_mapping():
    global _number_to_superscript
    if _number_to_superscript is not None:
        return  # 已经加载过，不重复读取

    wb = openpyxl.load_workbook("data.xlsx", data_only=True)
    ws = wb["遗漏偏差"]

    c_values = [ws[f"C{i}"].value for i in range(495, 501)]
    a_values = [ws[f"A{i}"].value for i in range(495, 501)]

    _number_to_superscript = {}
    for i, cell_value in enumerate(c_values):
        if isinstance(cell_value, str):
            for num in cell_value.split():
                _number_to_superscript[int(num)] = a_values[i]

def get_superscript(num):
    if _number_to_superscript is None:
        load_superscript_mapping()

    sup = _number_to_superscript.get(num)
    sup_str = _superscript_map.get(sup, " ") if sup is not None else " "
    return f"{num:2d}{sup_str.ljust(2)}"

            
_data = None  # 类变量，缓存整个 CSV 内容

@timeit
def load_data():
    global _data
    if _data is None:
        _data = pd.read_csv(r"scripts\ssq\ssq_history.csv", encoding='gbk', header=0)
    return _data

@timeit
def get_adjacent_numbers():
    # 读取 CSV 文件
    df = load_data()

    # 获取第一行的红球号码（索引 1 到 6）
    last_red_balls = df.iloc[0, 1:7].values

    # 计算相邻集合（去重）
    adjacent_set = set()
    for num in last_red_balls:
        adjacent_set.update({int(num) - 1, int(num), int(num) + 1})

    return adjacent_set

@timeit
def calculate_miss_count(front):
    df = load_data()
    # 读取最近 100 期历史记录
    df = df.head(100).iloc[::-1].reset_index(drop=True)

    # 转换数据格式，确保历史红球数据是整数
    history_numbers = [set(map(int, row[1:7])) for row in df.itertuples(index=False)]

    # 计算每个红球的遗漏次数
    miss_counts = []
    for num in front:
        miss_count = next((i for i, past in enumerate(reversed(history_numbers)) if num in past), 100)
        miss_counts.append(miss_count)
    
    return miss_counts
  
@timeit
def format_result(extra_randomness):
    df = load_data()
    base_qishu = df.iloc[0, 0]   # 第一行“期数”
    next_qishu = base_qishu + 1

    # 把结果和期数组装成列表
    row = list(extra_randomness) + [next_qishu]

    # 追加写入 ssq_predict.csv
    with open(r"scripts\ssq\ssq_predict.csv", "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(row)
        
    front = extra_randomness[:6]
    odd_count = sum(1 for num in front if num % 2 != 0)
    even_count = 6 - odd_count
    big_count = sum(1 for num in front if num > 16)
    small_count = 6 - big_count
    prime_count = sum(1 for num in front if is_prime(num))
    composite_count = 6 - prime_count
    sum_value = sum(front)
    adjacent_numbers = get_adjacent_numbers()
    adjacent_intersection = sorted(set(front) & adjacent_numbers)
    # 初始化 7 个区间，每个区间用 "_" 表示
    zones = ["_"] * 7

    # 计算区间分布（1-5 为区间 0，6-10 为区间 1，...，30-33 为区间 6）
    for num in front:
        zone_index = (num - 1) // 5  # 计算所属区间索引
        if zones[zone_index] == "_":
            zones[zone_index] = "1"  # 第一次出现，改为 "1"
        else:
            zones[zone_index] = str(int(zones[zone_index]) + 1)  # 出现次数+1
    # 计算平均遗漏值
    miss_counts = calculate_miss_count(front)
    # 计算平均遗漏值（保留 1 位小数）
    avg_miss_value = round(sum(miss_counts) / len(miss_counts), 1)
    formatted_numbers = " ".join(
        get_superscript(num) if i < 6 else f"+{num:2d}"  # 只对前6个数字处理
        for i, num in enumerate(extra_randomness)
    )
    sups = []
    for num in front:
        sup = _number_to_superscript.get(num)
        if sup is not None:
            sups.append(sup)
    ac = acValue(extra_randomness)
    return (f"{formatted_numbers}，"
        f"奇:偶 {odd_count}:{even_count}，"
        f"大:小 {big_count}:{small_count}，"
        f"质:合 {prime_count}:{composite_count}，"
        f"和值 {sum_value:3d}，"
        f"AC{ac:2d}，"
        f"区间 {''.join(zones)}，"
        f"遗漏{miss_counts}{avg_miss_value:4.1f}，"
        f"相邻数 {len(adjacent_intersection)}，"
        f"{sorted(sups)}。\n")
      
PRIMES = {2, 3, 5, 7, 11, 13, 17, 19, 23, 29, 31}

def is_prime(n):
    return n in PRIMES

@timeit
def filePrimeOrComposite(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    primeOrComposite(ws, 10)
    primeOrComposite(ws, 5)
    primeOrComposite(ws, 1)

def primeOrComposite(ws, sNo):
    prime = 0
    composite = 0
    for i in range(startSeral, startSeral + sNo):
        for j in range(0, 6):
            if (is_prime(redBalls[i][j])):
                prime = prime + 1
            else:
                composite = composite + 1

    if (sNo != 1):
        if (composite > prime):
            ws.append([f"近{sNo}期有", f"质 = {prime}", f"合 = {composite}", f"合：+{(composite * 100 // prime) - 100}%"])
        elif (composite < prime):
            ws.append([f"近{sNo}期有", f"质 = {prime}", f"合 = {composite}", f"质：+{(prime * 100 // composite) - 100}%"])
        else:
            ws.append([f"近{sNo}期有", f"质 = {prime}", f"合 = {composite}", '质 = 合'])
    else:
        ws.append(['本期比例:', '质:合 = %d:%d' % (prime, composite)])

"3.和值偏差"
@timeit
def sumOffset(ws):
    global arrDate, arrSeral
    sum = 0
    for i in range(0, 6):
        sum += redBalls[startSeral][i]
    cells = ['' for row in range(21)]
    cells[0] = arrDate[startSeral]
    cells[1] = int(arrSeral[startSeral])
    cells[2] = int(sum)
    cells[11] = '●'

    place = sum / 10
    # sum: cells[place + 1]  100: 11
    if place > 10:
        for i in range(12, int(place) + 1 + 1):
            cells[i] = '●'
    elif place < 10:
        for i in range(int(place) + 1, 11):
            cells[i] = '●'
    ws.append(cells)
    # 获取最大行数和最大列数
    max_row = ws.max_row
    max_column = ws.max_column

    ws['L%d' % max_row].font = fontRed

    # 遍历 D2 到 右下角 的所有单元格
    for row in ws.iter_rows(min_row=max_row, max_row=max_row, min_col=4, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def acValue(numbers):
    """
    计算前6个数的AC值。
    :param numbers: 一个包含至少6个整数的列表
    :return: AC值
    """
    if len(numbers) < 6:
        raise ValueError("输入的数组长度必须至少为6")

    numbers = sorted(numbers[:6])  # 取前6个数并排序
    difference_set = set()

    # 计算所有两两组合的绝对差值
    for i in range(len(numbers)):
        for j in range(i + 1, len(numbers)):
            difference_set.add(numbers[j] - numbers[i])

    D_t = len(difference_set)  # 不同差值的个数
    AC_value = D_t - (6 - 1)   # AC值计算

    return AC_value

@timeit
def fileAc(ws):
    global arrDate, arrSeral
    ac = acValue(redBalls[startSeral])
    cells = ['' for row in range(19)]
    cells[0] = arrDate[startSeral]
    cells[1] = int(arrSeral[startSeral])
    cells[2] = int(ac)
    cells[ac+3] = '●'
    ws.append(cells)

@timeit
def ColorAc(ws):
    # 提取 C2:C51 的值
    values = []
    for row in ws.iter_rows(min_row=2, max_row=51, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                values.append(cell.value)
    avg = round(sum(values) / len(values)) if values else 0
    # 偏移列，D列是第4列，从第4列开始偏移 avg
    target_col_idx = 4 + avg
    target_col_letter = get_column_letter(target_col_idx)
    # 设置字体样式（红色）
    red_font = Font(color="FF0000")

    # 填充目标列 E2:E51（或偏移后列）为 "●"，并设为红色字体
    for row in range(2, 52):
        cell = ws[f"{target_col_letter}{row}"]
        cell.value = "●"
        cell.font = red_font
        ac = ws[f"C{row}"].value
        a, b = sorted([avg, ac])
        if b - a > 1:
            for col in range(a+1, b):
                cn = get_column_letter(4+col)
                c = ws[f"{cn}{row}"]
                c.value = "●"
    # 获取最大行数和最大列数
    max_row = ws.max_row
    max_column = ws.max_column

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=4, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

"4.遗漏"
@timeit
def fileOmit(ws):
    global arrDate, arrSeral
    cells = ['' for row in range(7)]
    cells[0] = arrDate[startSeral]
    cells[1] = int(arrSeral[startSeral])
    for i in range(0, 6):
        cells[2] += '%4d' % (redBalls[startSeral][i])

    o = omit(startSeral)

    omitNum = 0 # 小于10的个数
    omitSum = 0 # 总和
    for i in range(0, 6):
        if o[i] < 10:
            omitNum = omitNum + 1
        cells[3] += '%4d' % (o[i])
        omitSum = omitSum + o[i]

    cells[4] = int(omitNum)
    cells[5] = int(omitSum)
    cells[6] = float('%.1f' % (omitSum / 6.0))
    ws.append(cells)

    rowNum = ws.max_row
    numCell = ws['E%d' % rowNum]
    if numCell.value == '6':
        numCell.font = fontRed

def omit(start):
    o = [-1 for col in range(6)] #遗漏次数
    for i in range(0, 6):
        for j in range(start + 1, 99):
            for k in range(0, 6):
                if redBalls[j][k] == redBalls[start][i]:
                    o[i] = j - start - 1
                    break
            if o[i] != -1:
                break
    return o

def count_red_ball_groups(redBalls, startSeral, group_size):
    num_groups = (33 + group_size - 1) // group_size  # 计算需要多少组
    bRange = [0] * num_groups  # 初始化计数数组

    for num in redBalls[startSeral]:  # 遍历该期红球号码
        group_index = (num - 1) // group_size  # 计算红球落在哪个组
        bRange[group_index] += 1  # 计数 +1

    return bRange

"5.区间"
@timeit
def ballRange(ws):
    global arrDate, arrSeral
    cells = ['' for row in range(26)]
    cells[0] = arrDate[startSeral]
    cells[1] = int(arrSeral[startSeral])

    bRange = count_red_ball_groups(redBalls, startSeral, 5)
    bRange_6 = count_red_ball_groups(redBalls, startSeral, 6)  # 按 6 个一组
    bRange_7 = count_red_ball_groups(redBalls, startSeral, 7)  # 按 7 个一组

    for i in range(0, 7):
        if bRange[i] == 0:
            cells[i+2] = '-'
        else:
            cells[i+2] = int(bRange[i])
    for i in range(0, len(bRange_6)):
        if bRange_6[i] == 0:
            cells[i+12] = '-'
        else:
            cells[i+12] = int(bRange_6[i])
    for i in range(0, len(bRange_7)):
        if bRange_7[i] == 0:
            cells[i+21] = '-'
        else:
            cells[i+21] = int(bRange_7[i])
    ws.append(cells)
    # 获取最大行数和最大列数
    max_row = ws.max_row
    max_column = ws.max_column

    for row in ws.iter_rows(min_row=max_row, max_row=max_row, min_col=2, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


"6.百分比, 热号，冷号，温号"
@timeit
def hotOrCold(ws):
    global arrDate, arrSeral
    nums = [0 for row in range(33)]
    cells = ['' for row in range(5)]
    cells[0] = arrDate[startSeral]
    cells[1] = arrSeral[startSeral]
    for i in range(startSeral, startSeral + 5):
        for j in range(0, 6):
            if nums[redBalls[i][j] - 1] == 0:
                nums[redBalls[i][j] - 1] = 1 #近1-5期出现
    for i in range(startSeral + 5, startSeral + 10):
        for j in range(0, 6):
            if nums[redBalls[i][j] - 1] == 0:
                nums[redBalls[i][j] - 1] = 2 #近6-10期出现
            elif nums[redBalls[i][j] -1] == 1:
                nums[redBalls[i][j] - 1] = 3 #两个区间都出现
    hot = ''
    cold = ''
    warm = ''
    hNum = 0
    cNum = 0
    wNum = 0
    for i in range(0, 33):
        if nums[i] == 0:
            cold += '%2d  ' % (i + 1)
            cNum = cNum + 1
            if (cNum >= 10):
                cNum = 0
                #cold += '\n'
        elif nums[i] == 3:
            hot += '%2d  ' % (i + 1)
            hNum = hNum + 1
            if (hNum >= 10):
                hNum = 0
                #hot += '\n'
        else:
            warm += '%2d  ' % (i + 1)
            wNum = wNum + 1
            if (wNum >= 10):
                wNum = 0
                #warm += '\n'

    cells[2] = hot
    cells[3] = cold
    cells[4] = warm
    ws.append(cells)


"7.遗漏偏差"
def omitDict(start, o):
    out = False
    for i in range(0, 6):
        for j in range(start + 1, 99):
            for k in range(0, 6):
                if redBalls[j][k] == redBalls[start][i]:
                    o[redBalls[start][i]] = j -start - 1
                    out = True
                    break
            if out:
                out = False
                break
    return o

@timeit
def omitOffset(ws, d):
    global arrDate, arrSeral

    redStr = ''
    for i in range(0, 6):
        redStr += '%2d ' % redBalls[startSeral][i]
    row_a = ws.max_row
    ws.append([f"近{d}期", '符合'])
    ws.append(['遗漏', '个数', '符合数字'])

    o = {}
    for i in range(startSeral + d - 1, startSeral - 1, -1):
        omitDict(i, o)
    arr = [[] for i in range(6)]
    cnt = {}
    for i in range(0, 6):
        for (k,v) in o.items():
            if v == i:
                arr[i].append(k)
        cnt[i] = len(arr[i])
    dictCnt = sorted(cnt.items(), key=lambda d:d[1], reverse = False)

    for i in range(0, 6):
        diStr = ''
        for j in range(1, 34):
            if omitTable[startSeral][j] == dictCnt[i][0]:
                diStr += '%2d  ' % (j)
        ws.append([dictCnt[i][0], dictCnt[i][1], diStr])
    row_b = ws.max_row
    for row in ws.iter_rows(min_row = row_a, max_row = row_b, min_col = 1, max_col = 2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

@timeit
def colorOmitOffset(ws, d):
    global arrDate, arrSeral, startSeral
    if startSeral >= 49:
        return

    startSeral = startSeral + 1
    redStr = ''
    rowNum = ws.max_row - 15
    o = {}
    for i in range(startSeral + d - 1, startSeral - 1, -1):
        omitDict(i, o)
    arr = [[] for i in range(6)]
    cnt = {}
    for i in range(0, 6):
        for (k,v) in o.items():
            if v == i:
                arr[i].append(k)
        cnt[i] = len(arr[i])
    dictCnt = sorted(cnt.items(), key=lambda d:d[1], reverse = False)

    for i in range(0, 6):
        redStr = ''
        for j in range(1, 34):
            if omitTable[startSeral][j] == dictCnt[i][0]:
                for cnt in range(0, 6):
                    if j == redBalls[startSeral - 1][cnt]:
                        redStr += '%2d ' % j

        ws['D%d' % rowNum].font = fontRed
        ws['D%d' % rowNum].value = redStr
        rowNum = rowNum + 1

    startSeral = startSeral -1

def fileOmitOffset(ws):
    global arrDate, arrSeral
    ws.append([''])
    ws.append([arrDate[startSeral] + arrSeral[startSeral]])
    omitOffset(ws, 5)
    colorOmitOffset(ws, 5)

    # omitOffset(ws, 6)
    #omitOffset(ws, 7)
    #omitOffset(ws, 8)
    #omitOffset(ws, 9)
    #omitOffset(ws, 10)


"(中)博彩趋势逆转"
def countOmit(ball, start):
    for i in range(start, 99):
        for j in range(0, 6):
            if redBalls[i][j] == ball:
                return i - start

# 当前颜色索引
color_index = 0

def next_font():
    """依次返回一个字体颜色（循环使用）"""
    global color_index
    color = colors[color_index % len(colors)]
    color_index += 1
    return Font(color=color, bold=True)

@timeit
def trendReverse(ws):
    global arrDate, arrSeral, color_index
    redStr = ''
    for i in range(0, 6):
        redStr += '%2d ' % redBalls[startSeral][i]
        last_row = ws.max_row
        offset = 5 # E列，每遇空值加一
        while last_row >= 1:
            cell = ws.cell(row = last_row, column = 2)
            if cell.value:
                if cell.value == redBalls[startSeral][i]:
                    target = ws.cell(row = last_row, column = offset)
                    target.value = cell.value
                    color_index = offset - 5
                    target.font = next_font()
            else:
                offset += 1
            last_row = last_row - 1

    ws.append([''])
    ws.append(['', arrDate[startSeral] + arrSeral[startSeral], '中奖号码: %s' % redStr])

    parn = ['' for row in range(6)]
    find = False
    num = 1

    # 筛选最近4期号码, 只查询上期中奖数字
    for i in range(0, 6):
        for j in range(startSeral + 4, startSeral, -1):
            for k in range(0, 6):
                if redBalls[j][k] == redBalls[startSeral][i]:
                    parn[i] += 'x'
                    find = True
                    num = 1
                    break
            if find == False:
                parn[i] += str(num)
                num = num + 1
            find = False
        parn[i] += 'x'
        num = 1

    find = False
    for i in range(0, 6):
        if parn[i] == '123xx':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om + 3 >= 17:
                ws.append(['博彩逆转：', redBalls[startSeral][i], ('模式：1..%d ' % (om + 3)) + parn[i][3:5].upper(), f"遗漏了{om+3}次"])
                find = True
        elif parn[i] == '12x1x':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om + 2 >= 17:
                ws.append(['博彩逆转：', redBalls[startSeral][i], ('模式：1..%d ' % (om + 2)) + parn[i][2:5].upper(), f"遗漏了{om+2}次"])
                find = True
        elif parn[i] == '1x12x':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om + 1 >= 17:
                ws.append(['博彩逆转：', redBalls[startSeral][i], ('模式：1..%d ' % (om + 1)) + parn[i][1:5].upper(), f"遗漏了{om+1}次"])
                find = True
        elif parn[i] == 'x123x':
            om = countOmit(redBalls[startSeral][i], startSeral + 5)
            if om >= 17:
                ws.append(['博彩逆转：', redBalls[startSeral][i], ('模式：1..%d ' % (om + 0)) + parn[i][0:5].upper(), f"遗漏了{om+0}次"])
                find = True

"层叠"
@timeit
def pile(ws):
    # 只查询中奖的号码
    for num in range(0, 6):
        ser = startSeral + 1
        ball = redBalls[startSeral][num]
        omit = omitTable[ser][ball]
        omit2 = 0
        omit3 = 0
        parn = []
        if omit >= 15:
            # x
            omit2 = omitTable[ser + omit + 1][ball]
            if omit2 == 3:
                # x123x
                parn.append('x123x 1..%d x' % omit)
                ser = ser + 4
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 2:
                # x12x
                parn.append('x12x 1..%d x' % omit)
                ser = ser + 3
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 1:
                # x1x
                parn.append('x1x 1..%d x' % omit)
                ser = ser + 2
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 0:
                # xx
                parn.append('xx 1..%d x' % omit)
                ser = ser + 1
                omit2 = omitTable[ser + omit + 1][ball]
                i = 0
                while omit2 == 0:
                    # xxx..
                    i = i + 1
                    ser = ser + 1
                    omit2 = omitTable[ser + omit + 1][ball]
                    parn.append('x')
            else:
                parn.append('x 1..%d x' % omit)

            if omit2 >= 8 and omit2 <= 15:
                # x
                omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                if omit3 == 3:
                    # x123x
                    ser = ser + 4
                    parn.append('x123x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 2:
                    # x12x
                    ser = ser + 3
                    parn.append('x12x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3== 1:
                    # x1x
                    ser = ser + 2
                    parn.append('x1x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 0:
                    # xx
                    ser = ser + 1
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                    parn.append('xx 1..%d ' % omit2)
                    i = 0
                    while omit3 == 0:
                        # xxx..
                        i = i + 1
                        ser = ser + 1
                        omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                        parn.append('x')
                else:
                    parn.append('x 1..%d ' % omit2)

                if omit3 >= 3 and omit3 <= 8:
                    parn.append('x 1..%d ' % omit3)
                    parn.reverse()
                    ws.append(['层叠：', ball , '模式：%s' % ''.join(parn).upper()])

"反向层叠"
@timeit
def rePile(ws):
    # 只查询中奖的号码
    for num in range(0, 6):
        ser = startSeral + 1
        ball = redBalls[startSeral][num]
        omit = omitTable[ser][ball]
        omit2 = 0
        omit3 = 0
        parn = []
        if omit <=8 and omit >= 3:
            # x
            omit2 = omitTable[ser + omit + 1][ball]
            if omit2 == 3:
                # x123x
                parn.append('x123x 1..%d x' % omit)
                ser = ser + 4
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 2:
                # x12x
                parn.append('x12x 1..%d x' % omit)
                ser = ser + 3
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 1:
                # x1x
                parn.append('x1x 1..%d x' % omit)
                ser = ser + 2
                omit2 = omitTable[ser + omit + 1][ball]
            elif omit2 == 0:
                # xx
                parn.append('xx 1..%d x' % omit)
                ser = ser + 1
                omit2 = omitTable[ser + omit + 1][ball]
                i = 0
                while omit2 == 0:
                    # xxx..
                    i = i + 1
                    ser = ser + 1
                    omit2 = omitTable[ser + omit + 1][ball]
                    parn.append('x')
            else:
                parn.append('x 1..%d x' % omit)

            if omit2 >= 8 and omit2 <= 15:
                # x
                omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                if omit3 == 3:
                    # x123x
                    ser = ser + 4
                    parn.append('x123x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 2:
                    # x12x
                    ser = ser + 3
                    parn.append('x12x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3== 1:
                    # x1x
                    ser = ser + 2
                    parn.append('x1x 1..%d ' % omit2)
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                elif omit3 == 0:
                    # xx
                    ser = ser + 1
                    omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                    parn.append('xx 1..%d ' % omit2)
                    i = 0
                    while omit3 == 0:
                        # xxx..
                        i = i + 1
                        ser = ser + 1
                        omit3 = omitTable[ser + omit + 1 + omit2 + 1][ball]
                        parn.append('x')
                else:
                    parn.append('x 1..%d ' % omit2)

                if omit3 >= 15:
                    parn.append('x 1..%d ' % omit3)
                    parn.reverse()
                    ws.append(['反向层叠：', ball , '模式：%s' % ''.join(parn).upper()])

"n底"
@timeit
def nBottom(ws):
    # 检验所有数字，预测提示
    for num in range(1, 34):
        ser = startSeral
        ball = num
        cnt = 1
        parn = []
        tRange = -1
        if omitTable[ser][ball] == 0:
            continue

        while True:
            omit = omitTable[ser][ball]
            i = 0
            while omit == 0:
                i = i + 1
                ser = ser + 1
                omit = omitTable[ser][ball]

            omit2 = omitTable[ser + omit + 1][ball]
            j = 0
            while omit2 == 0:
                j = j + 1
                ser = ser + 1
                omit2 = omitTable[ser + omit + 1][ball]

            if tRange == -1:
                tRange = omit

            if (omit >= 2 and omit2 >= 2) and (omit == tRange or omit == tRange + 1 or omit == tRange - 1) and (omit2 == tRange or omit2 == tRange + 1 or omit2 == tRange - 1):
                if cnt == 1:
                    for cnt in range(omit, 0, -1):
                        parn.append(str(cnt))

                while i != 0:
                    parn.append('x')
                    i = i - 1

                parn.append('x')
                cnt = cnt + 1
                ser = ser + omit + 1

                while j != 0:
                    parn.append('x')
                    j = j - 1

                for k in range(omit2, 0, -1):
                    parn.append(str(k))


            else:
                if cnt != 1:
                    parn.append('x')
                    parn.reverse()
                    parn = compress_ranges(parn)
                    ws.append(['%d 倍底：' % cnt, ball, '模式：%s' % ''.join(parn).upper()])
                    cnt = 1
                break

"旗式排列"
@timeit
def flagRange(ws):
    # 检验所有数字，预测提示
    # tips: 利用数组翻转
    for num in range(1, 34):
        ser = startSeral
        ball = num
        parn = []
        isOk = False
        #isOk2 = False
        if omitTable[ser][ball] == 0:
            continue
        isInit = True
        omit = 0
        cnt = 0
        while True:
            omit = omitTable[ser][ball]
            if isInit == True:
                if omit != 5 and omit != 6:
                    isOk = True
                    break

            if omit == 5 or omit == 6:
                ser = ser + omit + 1
                if isInit == False:
                    parn.append('x')
                for j in range(omit, 0, -1):
                    parn.append(str(j))
                if ser > 99:
                    break
                isInit = False
            elif omit < 3:
                # if omit == 3:
                #     #x123x
                #     parn.append('x123x')
                #     ser = ser + omit + 1
                if omit == 2:
                    #x12x
                    parn.append('x12x')
                    ser = ser + omit + 1
                elif omit == 1:
                    #x1x
                    parn.append('x1x')
                    ser = ser + omit + 1
                elif omit == 0:
                    #xx
                    parn.append('x')
                    while omit == 0:
                        ser = ser + 1
                        omit = omitTable[ser][ball]
                        parn.append('x')
                isInit = True
            else:
                break

            cnt = cnt + 1

        if cnt >= 2 and isOk == True:
            parn.reverse()
            parn = compress_ranges(parn)
            ws.append(['旗式排列：', ball, '模式：%s' % ''.join(parn).upper()])

def compress_ranges(parn):
    result = []
    buffer = []

    for item in parn:
        if re.fullmatch(r"\d+", item):  # 检测是否为数字
            buffer.append(item)
        else:
            if len(buffer) >= 4:
                if len(result) > 0:
                    result.append(" ")
                result.append(buffer[0])
                result.append("..")
                result.append(buffer[-1])
                result.append(" ")
            else:
                result.extend(buffer)
            result.append(item)
            buffer = []

    # 处理结尾的连续数字
    if len(buffer) >= 4:
        if len(result) > 0:
            result.append(" ")
        result.append(buffer[0])
        result.append("..")
        result.append(buffer[-1])
    else:
        result.extend(buffer)

    return result

@timeit
def generate_companion_matrix(ws):
    # 初始化 33x33 矩阵（从 1 到 33）
    companion_matrix = [[0] * 33 for _ in range(33)]

    # 读取 CSV 文件并解析红球号码
    with open(csv_file, newline='', encoding='gbk') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头

        for row in reader:
            red_balls = list(map(int, row[1:7]))  # 解析红球号码（跳过第一列期数）

            # 遍历所有红球的两两组合，并更新矩阵
            for i in range(len(red_balls)):
                for j in range(i + 1, len(red_balls)):  # 只处理上三角
                    num1, num2 = red_balls[i], red_balls[j]
                    companion_matrix[num1 - 1][num2 - 1] += 1  # (x, y) 位置加 1
                    companion_matrix[num2 - 1][num1 - 1] += 1  # (y, x) 位置加 1

    # 写入表头（1-33）
    ws.append([""] + list(range(1, 34)))

    # 写入伴侣数字矩阵
    for i in range(33):
        row_data = [i + 1] + companion_matrix[i]
        ws.append(row_data)

@timeit
def find_top_3num_combinations(top_n=10):
    counter = Counter()

    # 读取 CSV 文件
    with open(csv_file, newline='', encoding='gbk') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头

        for row in reader:
            red_balls = sorted(map(int, row[1:7]))  # 解析 & 排序 6 个红球

            # 生成所有三数组合，并计数
            for comb in combinations(red_balls, 3):
                counter[comb] += 1

    # 取出现次数最多的 top_n 个组合
    top_combinations = counter.most_common(top_n)

    # 输出结果
    print(f"出现次数最多的 {top_n} 组三数组合：")
    for (combo, count) in top_combinations:
        print(f"{combo}: {count} 次")

    return top_combinations

@timeit
def find_adjacent_number_pairs(top_n=10):
    counter = Counter()

    # 读取 CSV 文件
    with open(csv_file, newline='', encoding='gbk') as file:
        reader = csv.reader(file)
        next(reader)  # 跳过表头

        for row in reader:
            red_balls = sorted(map(int, row[1:7]))  # 解析并排序 6 个红球

            # 查找相邻数字对
            for i in range(len(red_balls) - 1):
                if red_balls[i] + 1 == red_balls[i + 1]:  # 检查是否是相邻数字
                    counter[(red_balls[i], red_balls[i + 1])] += 1

    # 取出现次数最多的 top_n 组
    top_adjacent_pairs = counter.most_common(top_n)

    # 输出结果
    print(f"出现次数最多的 {top_n} 组相邻数字对：")
    for (pair, count) in top_adjacent_pairs:
        print(f"{pair}: {count} 次")

    return top_adjacent_pairs

@timeit
def generate_miss_matrix(ws):
    df = pd.read_csv(csv_file, encoding='gbk', header=0)
    latest_50 = df.head(50).values[::-1]

    # 获取期数列表（第一列）
    periods = [str(int(row[0]) % 1000) for row in latest_50]

    # 生成表头
    ws.append(["数字"] + periods + ["遗漏", "中奖", "数字"])

    # 初始化1-33的中奖数据
    hit_counts = {num: 0 for num in range(1, 34)}

    for num in range(1, 34):
        row_temp = [num]
        row_data = [num]

        for record in reversed(latest_50):  # 逆序处理
            red_balls = set(map(int, record[1:7]))
            row_temp.append("X" if num in red_balls else "")
        # 计算遗漏次数（第一个"X"之前的空格数）
        miss_count = next((i for i, v in enumerate(row_temp[1:]) if v == "X"), len(row_temp) - 1)
        for record in latest_50:
            red_balls = set(map(int, record[1:7]))
            row_data.append("X" if num in red_balls else "")
            if num in red_balls:
                hit_counts[num] += 1

        # 追加遗漏次数、中奖次数和重复的数字
        row_data.extend([miss_count, hit_counts[num], num])
        ws.append(row_data)

    # 遍历 B2 到 右下角 的所有单元格
    for row in ws.iter_rows(min_row=2, max_row=34, min_col=2, max_col=51):
        for cell in reversed(row[1:]):  # 从第51列向左遍历
            if cell.value == "X":
                break  # 遇到 "X" 停止
            cell.value = "-"  # 填充中划线
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

@timeit
def generate_miss_pattern(ws):
    df = pd.read_csv(csv_file, encoding='gbk', header=0)
    data = df.head(1000).values[::-1]
    # 初始化遗漏统计
    miss_counts = {num: 0 for num in range(1, 34)}  # 记录当前遗漏次数
    blue_miss_counts = {num: 0 for num in range(1, 17)}
    occurrence = {num: collections.defaultdict(int) for num in range(1, 34)}  # 统计不同遗漏次数后中奖的次数
    blue_occurrence = {num: collections.defaultdict(int) for num in range(1, 17)}

    # **从最新一期（最后一行）开始向前遍历**
    for row in data:
        red_balls = set(map(int, row[1:7]))  # 提取6个红球号码
        blue_ball = int(row[7])  # 提取蓝球号码

        for num in range(1, 34):
            if num in red_balls:  # 该数字中奖
                occurrence[num][miss_counts[num]] += 1  # 记录该遗漏次数对应的中奖次数
                miss_counts[num] = 0  # 中奖后遗漏次数归零
            else:
                miss_counts[num] += 1  # 未中奖则遗漏次数增加

        # 处理蓝球（1-16）
        for num in range(1, 17):
            if num == blue_ball:  # 该蓝球中奖
                blue_occurrence[num][blue_miss_counts[num]] += 1  # 记录该遗漏次数对应的中奖次数
                blue_miss_counts[num] = 0  # 中奖后遗漏次数归零
            else:
                blue_miss_counts[num] += 1  # 未中奖则遗漏次数增加
    # 计算 n 的最大值（即最大的遗漏次数）
    max_miss = max(max(occ.keys(), default=0) for occ in occurrence.values())
    max_miss_blue = max(max(occ.keys(), default=0) for occ in blue_occurrence.values())
    # 生成表头
    ws.append(["数字", "遗漏"] + list(range(0, max_miss + 1)))

    # **提取最新一期的遗漏数据**
    for row_idx, num in enumerate(range(1, 34), start=2):
        row_data = [num, miss_counts[num]]  # 当前遗漏次数
        for i in range(0, max_miss + 1):
            row_data.append(occurrence[num][i] if occurrence[num][i] > 0 else "_")  # 0 替换为空字符
        ws.append(row_data)

        # **设置最新遗漏次数对应的单元格字体颜色为红**
        miss_count = miss_counts[num]  # 该数字的最新遗漏次数
        col_idx = miss_count + 3  # 对应的列索引 (Excel 索引从 1 开始，第4列才是 n=1)
        if col_idx <= max_miss + 3:  # 避免超出范围
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(color="FF0000")  # 设为红色
    # 遍历 C2 到 右下角 的所有单元格
    for row in ws.iter_rows(min_row=2, max_row=34, min_col=3, max_col=max_miss+3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # **插入一个空行**
    ws.append([])

    # **生成蓝球表头**
    ws.append(["蓝球", "遗漏"] + list(range(0, max_miss_blue + 1)))

    # **提取最新一期的蓝球遗漏数据**
    for row_idx, num in enumerate(range(1, 17), start=ws.max_row + 1):
        row_data = [num, blue_miss_counts[num]]  # 当前遗漏次数
        for i in range(0, max_miss_blue + 1):
            row_data.append(blue_occurrence[num][i] if blue_occurrence[num][i] > 0 else "_")
        ws.append(row_data)

        # **设置蓝球最新遗漏次数对应的单元格字体颜色为红**
        miss_count = blue_miss_counts[num]
        col_idx = miss_count + 3
        if col_idx <= max_miss_blue + 3:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(color="FF0000")

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 遍历 C36 到 右下角 的所有单元格
    for row in ws.iter_rows(min_row=35, max_row=52, min_col=3, max_col=max_miss_blue+3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def calc_cell_width(value):
    """ 计算单元格内容的最佳宽度 """
    value = str(value) if value is not None else ""
    chinese_count = len(re.findall(r'[\u4e00-\u9fff]', value))  # 统计中文字符个数
    other_count = len(value) - chinese_count  # 统计非中文字符个数
    return chinese_count * 2 + other_count  # 中文算2个宽度，其他算1个宽度

@timeit
def auto_adjust_column_width(wb):
    """遍历工作簿中的所有 Sheet，并自动调整列宽"""
    for ws in wb.worksheets:  # 遍历所有工作表
        for col in ws.columns:  # 遍历每一列
            max_length = 0
            col_letter = col[0].column_letter  # 获取列字母 (A, B, C, ...)

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, calc_cell_width(cell.value))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2  # 适当增加宽度

        max_row = ws.max_row
        # 设置活动单元格为 A 列的最大行数对应的单元格
        selection_length = len(ws.sheet_view.selection)
        ws.sheet_view.selection[selection_length-1].activeCell = f"A{max_row}"
        ws.sheet_view.selection[selection_length-1].sqref = f"A{max_row}"

def is_number_with_spaces(value):
    if isinstance(value, (int, float)):  # 纯数字
        return True
    if isinstance(value, str):  # 字符串情况
        return bool(re.fullmatch(r'\s*\d+(\s+\d+)*\s*', value))
    return False  # 其他类型都返回 False

@timeit
def set_workbook_font(wb, font_name="IBM Plex Sans SC", num_font="IBM Plex Mono", font_size=9):
    """遍历工作簿中的所有 Sheet，并为所有单元格修改字体和字号，保留原有颜色"""
    for ws in wb.worksheets:  # 遍历所有 Sheet
        for row in ws.iter_rows():  # 遍历所有行
            for cell in row:  # 遍历该行的所有单元格
                if cell.value is not None:  # 只设置非空单元格
                    old_font = cell.font  # 获取原有字体
                    if is_number_with_spaces(cell.value):
                        new_font = Font(
                            name=num_font,
                            size=font_size,
                            bold=old_font.bold,
                            italic=old_font.italic,
                            color=old_font.color,
                            underline=old_font.underline,
                            strike=old_font.strike
                        )
                    else:
                        new_font = Font(
                            name=font_name,
                            size=font_size,
                            bold=old_font.bold,
                            italic=old_font.italic,
                            color=old_font.color,
                            underline=old_font.underline,
                            strike=old_font.strike
                        )
                    cell.font = new_font  # 只修改字体和字号，保留颜色等其他样式
                    alignment = copy(cell.alignment)
                    alignment.vertical = 'center'
                    cell.alignment = alignment

@timeit
def merge(wb):
    src_ws = wb["大小数"]    # 源 sheet
    dst_ws = wb["奇偶数"]  # 目标 sheet

    for row in range(1, 251):
        for col in range(2, 5):  # B~D 是第 2~4 列
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row, column=col + 6)  # G~I 是第 7~9 列

            # 复制值
            dst_cell.value = src_cell.value

            # 可选：复制样式、字体等（如果需要）
            if src_cell.has_style:
                dst_cell.font = src_cell.font
                dst_cell.border = src_cell.border
                dst_cell.fill = src_cell.fill
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = src_cell.protection
                dst_cell.alignment = src_cell.alignment
    src_ws.sheet_state = 'hidden'  # 设置为隐藏

    src_ws = wb["质合数"]    # 源 sheet

    for row in range(1, 251):
        for col in range(2, 5):  # B~D 是第 2~4 列
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row, column=col + 12)

            # 复制值
            dst_cell.value = src_cell.value

            # 可选：复制样式、字体等（如果需要）
            if src_cell.has_style:
                dst_cell.font = src_cell.font
                dst_cell.border = src_cell.border
                dst_cell.fill = src_cell.fill
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = src_cell.protection
                dst_cell.alignment = src_cell.alignment
    src_ws.sheet_state = 'hidden'  # 设置为隐藏

    src_ws = wb["AC值"]    # 源 sheet
    dst_ws = wb["和值偏差"]  # 目标 sheet

    for row in range(1, 52):
        for col in range(3, 20):
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row, column=col + 24)

            # 复制值
            dst_cell.value = src_cell.value

            # 可选：复制样式、字体等（如果需要）
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)
    src_ws.sheet_state = 'hidden'  # 设置为隐藏

def createXlsx():
    wb = Workbook()

    ws0 = wb.active
    ws0.title = '文件信息'
    ws0['A1'] = '最新期数'

    ws1 = wb.create_sheet(str(1))
    ws1.title = '奇偶数'

    ws2 = wb.create_sheet(str(2))
    ws2.title = '大小数'

    ws3 = wb.create_sheet(str(3))
    ws3.title = '质合数'

    ws4 = wb.create_sheet(str(4))
    ws4.title = '和值偏差'
    ws4.append(['日期', '期数', '和值'] + [f"{i}" for i in range(20, 200, 10)])

    ws5 = wb.create_sheet(str(5))
    ws5.title = 'AC值'
    ws5.append(['日期', '期数', 'AC值'] + [f"{i}" for i in range(0, 16, 1)])

    ws6 = wb.create_sheet(str(6))
    ws6.title = '区间'
    ws6.append([
        '日期', '期数',
        '1-5', '6-10', '11-15', '16-20', '21-25', '26-30', '31-33', '', '', '',
        '1-6', '7-12', '13-18', '19-24', '25-30', '31-33', '', '', '',
        '1-7', '8-14', '15-21', '22-28', '29-33'
    ])

    ws7 = wb.create_sheet(str(7))
    ws7.title = '百分比'
    ws7.append(['日期', '期数', '热号', '冷号', '温号'])

    ws8 = wb.create_sheet(str(8))
    ws8.title = '遗漏'
    ws8.append(['日期', '期数', '中奖号码', '遗漏情况', '遗漏少于10次的个数', '总计', '平均'])

    ws9 = wb.create_sheet(str(9))
    ws9.title = '遗漏偏差'

    wsa = wb.create_sheet(str(10))
    wsa.title = '遗漏模式'

    wsb = wb.create_sheet(str(11))
    wsb.title = '中期表'

    wsc = wb.create_sheet(str(12))
    wsc.title = '小型期数'

    wsd = wb.create_sheet(str(13))
    wsd.title = '伴侣数字'

    return wb


"完整性检查"
def checkComplete(ws):
    global arrDate, arrSeral, startSeral
    v = ws['A2'].value
    if v == arrSeral[0]:
        print("no need to update.")
        exit()
    elif v == arrSeral[1]:
        startSeral = 0
        print("update the newest.")
        return True
    else:
        for i in range(2, 100):
            if v == arrSeral[i]:
                startSeral = i - 1 # 从下一期开始更新
                #print 'update from %s' % arrSeral[startSeral]
                break
        return False


def addInfo(ws):
    ws['A2'] = arrSeral[0]
    ws.sheet_state = 'hidden'

def countBall():
    fileOddOrEven(wb['奇偶数'])
    fileBigOrSmall(wb['大小数'])
    filePrimeOrComposite(wb['质合数'])
    sumOffset(wb['和值偏差'])
    fileAc(wb['AC值'])
    ballRange(wb['区间'])
    hotOrCold(wb['百分比'])
    fileOmit(wb['遗漏'])
    fileOmitOffset(wb['遗漏偏差'])
    trendReverse(wb['中期表'])
    pile(wb['中期表'])
    rePile(wb['中期表'])
    nBottom(wb['中期表'])
    flagRange(wb['中期表'])

if __name__ == '__main__':
    importlib.reload(sys)

    exist = os.path.isfile('./data.xlsx')
    if exist:
        os.remove('./data.xlsx')
    exist = os.path.isfile('./data.xlsx')
    getData() #近100期
    if exist:
        wb = load_workbook('./data.xlsx')
        if checkComplete(wb['文件信息']):
            countBall()
        else:
            while (startSeral != -1):
                countBall()
                startSeral = startSeral - 1
    else:
        wb = createXlsx()
        startSeral = 49
        while (startSeral != -1):
            countBall()
            startSeral = startSeral - 1

    generate_companion_matrix(wb['伴侣数字'])
    generate_miss_matrix(wb['小型期数'])
    generate_miss_pattern(wb['遗漏模式'])
    ColorAc(wb['AC值'])
    merge(wb)
    addInfo(wb['文件信息'])
        # 需要设置的 Sheet 名称列表
    target_sheets = ['和值偏差', 'AC值', '区间', '百分比', '遗漏', '伴侣数字', '小型期数', '遗漏模式']

    # 遍历指定的 Sheet 进行设置
    for sheet_name in target_sheets:
        if sheet_name in wb.sheetnames:  # 确保 Sheet 存在
            sheet = wb[sheet_name]
            sheet.freeze_panes = 'C2'  # 冻结首行和前两列，需要设置为 第一个未被冻结的单元格
            if sheet_name == '小型期数' or sheet_name == '伴侣数字':
                sheet.freeze_panes = 'B2'

            for col in sheet.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    auto_adjust_column_width(wb)
    set_workbook_font(wb, font_name="IBM Plex Sans SC", num_font="IBM Plex Mono", font_size=9)
    wb.save('./data.xlsx')
    # analyze_lottery_repeats()
    # find_top_3num_combinations()
    # find_adjacent_number_pairs()